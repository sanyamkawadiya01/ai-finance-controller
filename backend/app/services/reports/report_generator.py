import io
import datetime
from typing import List, Dict, Any, Optional

from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.app.schemas.schemas import (
    ReconciliationResultSchema,
    DashboardSummarySchema,
    EvaluationReportSchema,
    InvoiceSchema,
    BankTransactionSchema
)


class NumberedCanvas(canvas.Canvas):
    """
    Two-pass canvas to dynamically compute and print total page count
    and professional header/footer rules.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#64748B"))
        
        # Top Header (pages > 1)
        if self._pageNumber > 1:
            self.drawString(36, 576, "AI Finance Controller — Payment Reconciliation Report")
            self.drawRightString(756, 576, datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
            self.setStrokeColor(colors.HexColor("#CBD5E1"))
            self.setLineWidth(0.5)
            self.line(36, 568, 756, 568)
            
        # Bottom Footer (all pages)
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(756, 20, page_text)
        self.drawString(36, 20, "Confidential — AI Finance Controller System Report")
        self.setStrokeColor(colors.HexColor("#CBD5E1"))
        self.setLineWidth(0.5)
        self.line(36, 32, 756, 32)
        self.restoreState()


class PDFReportGenerator:
    @staticmethod
    def generate(
        summary: DashboardSummarySchema,
        results: List[ReconciliationResultSchema],
        invoices: List[InvoiceSchema],
        transactions: List[BankTransactionSchema],
        evaluation: Optional[EvaluationReportSchema],
        dataset_source: str = "default"
    ) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            leftMargin=36,
            rightMargin=36,
            topMargin=45,
            bottomMargin=45
        )

        styles = getSampleStyleSheet()
        
        # Custom Typography Styles
        title_style = ParagraphStyle(
            'DocTitle',
            parent=styles['Title'],
            fontName='Helvetica-Bold',
            fontSize=22,
            leading=26,
            textColor=colors.HexColor("#0F172A"),
            alignment=0
        )
        
        subtitle_style = ParagraphStyle(
            'DocSubtitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#475569")
        )

        section_heading = ParagraphStyle(
            'SectionHeading',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=13,
            leading=16,
            textColor=colors.HexColor("#1E293B"),
            spaceBefore=12,
            spaceAfter=6
        )

        cell_text = ParagraphStyle(
            'CellText',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=7.5,
            leading=9,
            textColor=colors.HexColor("#1E293B")
        )

        cell_text_bold = ParagraphStyle(
            'CellTextBold',
            parent=cell_text,
            fontName='Helvetica-Bold'
        )

        cell_text_header = ParagraphStyle(
            'CellHeader',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            leading=10,
            textColor=colors.white
        )

        story = []
        gen_time = datetime.datetime.now().strftime("%B %d, %Y at %I:%M %p")

        # ---------------------------------------------------------
        # 1. REPORT HEADER
        # ---------------------------------------------------------
        header_data = [
            [
                Paragraph("<b>AI Finance Controller</b><br/><font size=16 color='#2563EB'><b>Payment Reconciliation Report</b></font>", title_style),
                Paragraph(f"<b>Generated:</b> {gen_time}<br/>"
                          f"<b>Dataset Source:</b> {dataset_source.capitalize()}<br/>"
                          f"<b>Total Invoices:</b> {summary.total_invoices} | <b>Total Bank Txns:</b> {summary.total_transactions}", subtitle_style)
            ]
        ]
        header_table = Table(header_data, colWidths=[420, 300])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (1,0), (1,0), 'RIGHT')
        ]))
        story.append(header_table)
        story.append(Spacer(1, 10))
        story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#2563EB"), spaceBefore=2, spaceAfter=12))

        # ---------------------------------------------------------
        # 2. EXECUTIVE SUMMARY
        # ---------------------------------------------------------
        story.append(Paragraph("Executive Summary & Reconciliation Metrics", section_heading))
        
        match_rate = (summary.matched_count / summary.total_transactions * 100) if summary.total_transactions > 0 else 0.0

        kpi_data = [
            [
                Paragraph("Total Invoices", cell_text_bold),
                Paragraph("Total Transactions", cell_text_bold),
                Paragraph("Auto Matched", cell_text_bold),
                Paragraph("Needs Review", cell_text_bold),
                Paragraph("Unmatched", cell_text_bold),
                Paragraph("Duplicates", cell_text_bold),
                Paragraph("Late Payments", cell_text_bold),
                Paragraph("Overdue Invoices", cell_text_bold),
                Paragraph("Overall Match Rate", cell_text_bold)
            ],
            [
                Paragraph(str(summary.total_invoices), cell_text),
                Paragraph(str(summary.total_transactions), cell_text),
                Paragraph(f"<font color='#16A34A'><b>{summary.matched_count}</b></font>", cell_text),
                Paragraph(f"<font color='#D97706'><b>{summary.human_review_count}</b></font>", cell_text),
                Paragraph(f"<font color='#DC2626'><b>{summary.unmatched_count}</b></font>", cell_text),
                Paragraph(f"<font color='#9333EA'><b>{summary.duplicate_count}</b></font>", cell_text),
                Paragraph(f"<font color='#EA580C'><b>{summary.late_payments_count}</b></font>", cell_text),
                Paragraph(f"<font color='#B91C1C'><b>{summary.overdue_invoices_count}</b></font>", cell_text),
                Paragraph(f"<b>{match_rate:.1f}%</b>", cell_text)
            ]
        ]
        
        kpi_table = Table(kpi_data, colWidths=[80, 80, 80, 80, 80, 80, 80, 80, 80])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#F1F5F9")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor("#0F172A")),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 14))

        # Build lookup dict for invoice amounts
        inv_amt_map = {inv.invoice_id: inv.invoice_amount for inv in invoices}

        # ---------------------------------------------------------
        # 3. RECONCILIATION RESULTS TABLE
        # ---------------------------------------------------------
        story.append(Paragraph("Detailed Reconciliation Results", section_heading))
        
        rec_headers = [
            Paragraph("Txn ID", cell_text_header),
            Paragraph("Date", cell_text_header),
            Paragraph("Customer Name", cell_text_header),
            Paragraph("Amount", cell_text_header),
            Paragraph("Curr", cell_text_header),
            Paragraph("Matched Inv", cell_text_header),
            Paragraph("Inv Amt", cell_text_header),
            Paragraph("Match Type", cell_text_header),
            Paragraph("Conf", cell_text_header),
            Paragraph("Level", cell_text_header),
            Paragraph("Status", cell_text_header),
            Paragraph("Late", cell_text_header),
            Paragraph("Dup", cell_text_header),
            Paragraph("Action", cell_text_header)
        ]
        
        table_rows = [rec_headers]

        for r in results:
            inv_amt_str = f"${inv_amt_map[r.invoice_id]:,.2f}" if (r.invoice_id and r.invoice_id in inv_amt_map) else "—"
            conf_str = f"{r.confidence * 100:.0f}%"
            dup_str = "Yes" if r.is_duplicate else "No"
            days_late_str = str(r.days_late) if r.days_late > 0 else "0"

            # Color coding for status & level
            level_color = "#16A34A" if r.confidence_level == "HIGH" else ("#D97706" if r.confidence_level == "MEDIUM" else "#DC2626")
            status_color = "#16A34A" if r.payment_status == "on_time" else ("#EA580C" if r.payment_status == "late" else "#DC2626")
            
            row = [
                Paragraph(r.transaction_id, cell_text_bold),
                Paragraph(r.transaction_date, cell_text),
                Paragraph(r.customer_name[:20], cell_text),
                Paragraph(f"${r.amount:,.2f}", cell_text),
                Paragraph(r.currency, cell_text),
                Paragraph(r.invoice_id or "—", cell_text),
                Paragraph(inv_amt_str, cell_text),
                Paragraph(r.match_type.capitalize(), cell_text),
                Paragraph(conf_str, cell_text),
                Paragraph(f"<font color='{level_color}'><b>{r.confidence_level}</b></font>", cell_text),
                Paragraph(f"<font color='{status_color}'><b>{r.payment_status.replace('_',' ')}</b></font>", cell_text),
                Paragraph(days_late_str, cell_text),
                Paragraph(f"<font color='{ '#9333EA' if r.is_duplicate else '#64748B' }'><b>{dup_str}</b></font>", cell_text),
                Paragraph(r.action, cell_text_bold)
            ]
            table_rows.append(row)

        # Col widths total = 720pt (exact landscape printable width)
        col_widths = [55, 52, 85, 55, 30, 65, 55, 52, 32, 42, 45, 32, 30, 90]
        results_table = Table(table_rows, colWidths=col_widths, repeatRows=1)
        
        t_style = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1E293B")),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor("#E2E8F0")),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]
        
        # Alternating row colors
        for i in range(1, len(table_rows)):
            bg = colors.HexColor("#F8FAFC") if i % 2 == 1 else colors.white
            t_style.append(('BACKGROUND', (0, i), (-1, i), bg))
            
        results_table.setStyle(TableStyle(t_style))
        story.append(results_table)
        story.append(Spacer(1, 14))

        # ---------------------------------------------------------
        # 4. EXCEPTIONS / ATTENTION REQUIRED
        # ---------------------------------------------------------
        exceptions = [r for r in results if r.match_type == "unmatched" or r.action == "HUMAN_REVIEW" or r.is_duplicate or r.payment_status in ("late", "overdue")]
        
        story.append(Paragraph("Exceptions & Attention Required", section_heading))
        if not exceptions:
            story.append(Paragraph("<font color='#16A34A'><b>No exceptions or flag issues detected in current reconciliation run.</b></font>", subtitle_style))
        else:
            exc_headers = [
                Paragraph("Txn ID", cell_text_header),
                Paragraph("Customer", cell_text_header),
                Paragraph("Amount", cell_text_header),
                Paragraph("Matched Inv", cell_text_header),
                Paragraph("Issue Flag", cell_text_header),
                Paragraph("Payment Status", cell_text_header),
                Paragraph("Recommended Action", cell_text_header)
            ]
            exc_rows = [exc_headers]

            for r in exceptions:
                flags = []
                if r.match_type == "unmatched":
                    flags.append("Unmatched Txn")
                if r.action == "HUMAN_REVIEW":
                    flags.append("Needs Review")
                if r.is_duplicate:
                    flags.append("Duplicate Payment")
                if r.payment_status == "late":
                    flags.append(f"Late ({r.days_late}d)")
                if r.payment_status == "overdue":
                    flags.append("Overdue")

                flag_str = ", ".join(flags)
                
                exc_rows.append([
                    Paragraph(r.transaction_id, cell_text_bold),
                    Paragraph(r.customer_name[:25], cell_text),
                    Paragraph(f"${r.amount:,.2f} {r.currency}", cell_text),
                    Paragraph(r.invoice_id or "Unassigned", cell_text),
                    Paragraph(f"<font color='#DC2626'><b>{flag_str}</b></font>", cell_text),
                    Paragraph(r.payment_status.replace("_", " "), cell_text),
                    Paragraph(r.action, cell_text_bold)
                ])

            exc_table = Table(exc_rows, colWidths=[70, 110, 80, 80, 150, 80, 150], repeatRows=1)
            exc_style = [
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#7F1D1D")),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor("#FCA5A5")),
                ('TOPPADDING', (0,0), (-1,-1), 4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ]
            for i in range(1, len(exc_rows)):
                bg = colors.HexColor("#FEF2F2") if i % 2 == 1 else colors.white
                exc_style.append(('BACKGROUND', (0, i), (-1, i), bg))
            exc_table.setStyle(TableStyle(exc_style))
            story.append(exc_table)
            
        story.append(Spacer(1, 14))

        # ---------------------------------------------------------
        # 5. EVALUATION BENCHMARK SECTION (If Ground Truth Exists)
        # ---------------------------------------------------------
        story.append(Paragraph("Ground Truth Evaluation Benchmark", section_heading))
        if evaluation is None:
            story.append(Paragraph("<i>Ground truth evaluation is unavailable for this run (no ground_truth.csv uploaded).</i>", subtitle_style))
        else:
            eval_data = [
                [
                    Paragraph("Overall Accuracy", cell_text_bold),
                    Paragraph("Overall Precision", cell_text_bold),
                    Paragraph("Overall Recall", cell_text_bold),
                    Paragraph("Overall F1 Score", cell_text_bold)
                ],
                [
                    Paragraph(f"<b>{evaluation.overall_accuracy * 100:.1f}%</b>", cell_text),
                    Paragraph(f"<b>{evaluation.overall_precision * 100:.1f}%</b>", cell_text),
                    Paragraph(f"<b>{evaluation.overall_recall * 100:.1f}%</b>", cell_text),
                    Paragraph(f"<font color='#2563EB'><b>{evaluation.overall_f1 * 100:.1f}%</b></font>", cell_text)
                ]
            ]
            eval_table = Table(eval_data, colWidths=[180, 180, 180, 180])
            eval_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#EFF6FF")),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#BFDBFE")),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ]))
            story.append(eval_table)
            story.append(Spacer(1, 8))

            # Confusion Matrix Table
            if evaluation.confusion_matrix:
                story.append(Paragraph("<b>Confusion Matrix (Predicted vs Actual)</b>", subtitle_style))
                cm_keys = list(evaluation.confusion_matrix.keys())
                cm_header = [Paragraph("Actual \\ Predicted", cell_text_bold)] + [Paragraph(k, cell_text_bold) for k in cm_keys]
                cm_rows = [cm_header]
                for actual_k in cm_keys:
                    row = [Paragraph(actual_k, cell_text_bold)]
                    for pred_k in cm_keys:
                        val = evaluation.confusion_matrix.get(actual_k, {}).get(pred_k, 0)
                        cell_color = "#16A34A" if (actual_k == pred_k and val > 0) else "#1E293B"
                        row.append(Paragraph(f"<font color='{cell_color}'>{val}</font>", cell_text))
                    cm_rows.append(row)

                cm_col_width = min(120, int(720 / (len(cm_keys) + 1)))
                cm_table = Table(cm_rows, colWidths=[cm_col_width] * (len(cm_keys) + 1))
                cm_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#F8FAFC")),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor("#CBD5E1")),
                    ('TOPPADDING', (0,0), (-1,-1), 4),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                ]))
                story.append(cm_table)

        doc.build(story, canvasmaker=NumberedCanvas)
        return buffer.getvalue()


class ExcelReportGenerator:
    @staticmethod
    def generate(
        summary: DashboardSummarySchema,
        results: List[ReconciliationResultSchema],
        invoices: List[InvoiceSchema],
        transactions: List[BankTransactionSchema],
        evaluation: Optional[EvaluationReportSchema],
        dataset_source: str = "default"
    ) -> bytes:
        wb = openpyxl.Workbook()
        
        # Styles
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        title_font = Font(name="Calibri", size=16, bold=True, color="0F172A")
        section_font = Font(name="Calibri", size=13, bold=True, color="1E293B")
        bold_font = Font(name="Calibri", size=11, bold=True, color="0F172A")
        regular_font = Font(name="Calibri", size=11, color="1E293B")
        
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        kpi_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        # ---------------------------------------------------------
        # SHEET 1 — SUMMARY
        # ---------------------------------------------------------
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.views.sheetView[0].showGridLines = True

        ws_summary["A1"] = "AI Finance Controller — Payment Reconciliation Summary"
        ws_summary["A1"].font = title_font

        ws_summary["A2"] = f"Report Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')} | Dataset Source: {dataset_source.capitalize()}"
        ws_summary["A2"].font = Font(name="Calibri", size=10, italic=True, color="64748B")

        # KPI Block
        ws_summary["A4"] = "Executive Key Performance Indicators"
        ws_summary["A4"].font = section_font

        kpi_headers = [
            "Total Invoices", "Total Transactions", "Total Invoiced ($)", "Total Received ($)",
            "Auto Matched", "Needs Review", "Unmatched", "Duplicates",
            "Late Payments", "Overdue Invoices", "Overall Match Rate"
        ]
        
        match_rate = (summary.matched_count / summary.total_transactions) if summary.total_transactions > 0 else 0.0

        kpi_values = [
            summary.total_invoices,
            summary.total_transactions,
            summary.total_amount_invoiced,
            summary.total_amount_received,
            summary.matched_count,
            summary.human_review_count,
            summary.unmatched_count,
            summary.duplicate_count,
            summary.late_payments_count,
            summary.overdue_invoices_count,
            match_rate
        ]

        for col_num, (h, v) in enumerate(zip(kpi_headers, kpi_values), start=1):
            cell_h = ws_summary.cell(row=5, column=col_num, value=h)
            cell_h.font = header_font
            cell_h.fill = header_fill
            cell_h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            cell_v = ws_summary.cell(row=6, column=col_num, value=v)
            cell_v.font = bold_font
            cell_v.fill = kpi_fill
            cell_v.border = thin_border
            cell_v.alignment = Alignment(horizontal="center", vertical="center")
            
            if "($)" in h:
                cell_v.number_format = '$#,##0.00'
            elif "Rate" in h:
                cell_v.number_format = '0.0%'

        # Payment Status Breakdown Table
        ws_summary["A9"] = "Payment Status & Exceptions Breakdown"
        ws_summary["A9"].font = section_font

        status_headers = ["Payment Category", "Count", "Percentage of Total Txns"]
        for c, h in enumerate(status_headers, 1):
            cell = ws_summary.cell(row=10, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

        total_txns = max(1, summary.total_transactions)
        status_rows = [
            ("Auto Matched (High Confidence)", summary.matched_count, summary.matched_count / total_txns),
            ("Needs Review (Medium Confidence)", summary.human_review_count, summary.human_review_count / total_txns),
            ("Unmatched Transactions", summary.unmatched_count, summary.unmatched_count / total_txns),
            ("Duplicate Payment Attempts", summary.duplicate_count, summary.duplicate_count / total_txns),
            ("Late Payments", summary.late_payments_count, summary.late_payments_count / total_txns),
            ("Overdue Invoices", summary.overdue_invoices_count, summary.overdue_invoices_count / max(1, summary.total_invoices))
        ]

        for r_idx, (cat, cnt, pct) in enumerate(status_rows, start=11):
            c1 = ws_summary.cell(row=r_idx, column=1, value=cat)
            c2 = ws_summary.cell(row=r_idx, column=2, value=cnt)
            c3 = ws_summary.cell(row=r_idx, column=3, value=pct)
            
            c1.font = regular_font
            c2.font = bold_font
            c3.font = regular_font
            c3.number_format = '0.0%'
            
            for cell in (c1, c2, c3):
                cell.border = thin_border

        # Evaluation Metrics Section (if exists)
        if evaluation:
            ws_summary["A19"] = "Ground Truth Benchmark Evaluation Metrics"
            ws_summary["A19"].font = section_font

            eval_headers = ["Metric Name", "Score"]
            for c, h in enumerate(eval_headers, 1):
                cell = ws_summary.cell(row=20, column=c, value=h)
                cell.font = header_font
                cell.fill = header_fill

            eval_metrics = [
                ("Overall Accuracy", evaluation.overall_accuracy),
                ("Overall Precision", evaluation.overall_precision),
                ("Overall Recall", evaluation.overall_recall),
                ("Overall F1 Score", evaluation.overall_f1)
            ]

            for r_idx, (m_name, m_val) in enumerate(eval_metrics, start=21):
                c1 = ws_summary.cell(row=r_idx, column=1, value=m_name)
                c2 = ws_summary.cell(row=r_idx, column=2, value=m_val)
                c1.font = regular_font
                c2.font = bold_font
                c2.number_format = '0.0%'
                c1.border = thin_border
                c2.border = thin_border

        # ---------------------------------------------------------
        # SHEET 2 — RECONCILIATION RESULTS
        # ---------------------------------------------------------
        ws_rec = wb.create_sheet(title="Reconciliation Results")
        ws_rec.views.sheetView[0].showGridLines = True
        
        inv_amt_map = {inv.invoice_id: inv.invoice_amount for inv in invoices}
        
        rec_headers = [
            "Transaction ID", "Transaction Date", "Customer Name", "Amount", "Currency",
            "Matched Invoice ID", "Invoice Amount", "Match Type", "Confidence",
            "Confidence Level", "Payment Status", "Days Late", "Is Duplicate", "Action", "Reasons / Notes"
        ]
        
        for col, h in enumerate(rec_headers, 1):
            cell = ws_rec.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for row_idx, r in enumerate(results, start=2):
            inv_amt = inv_amt_map.get(r.invoice_id, None) if r.invoice_id else None
            reasons_str = "; ".join(r.reasons) if r.reasons else ""

            row_data = [
                r.transaction_id,
                r.transaction_date,
                r.customer_name,
                r.amount,
                r.currency,
                r.invoice_id or "",
                inv_amt,
                r.match_type,
                r.confidence,
                r.confidence_level,
                r.payment_status,
                r.days_late,
                "TRUE" if r.is_duplicate else "FALSE",
                r.action,
                reasons_str
            ]

            for col_idx, val in enumerate(row_data, start=1):
                cell = ws_rec.cell(row=row_idx, column=col_idx, value=val)
                cell.font = regular_font
                cell.border = thin_border
                
                if col_idx in (4, 7) and isinstance(val, (int, float)):
                    cell.number_format = '$#,##0.00'
                elif col_idx == 9 and isinstance(val, (int, float)):
                    cell.number_format = '0.0%'
                elif col_idx in (1, 2, 5, 8, 10, 11, 12, 13, 14):
                    cell.alignment = Alignment(horizontal="center")

        ws_rec.freeze_panes = 'A2'
        ws_rec.auto_filter.ref = ws_rec.dimensions

        # ---------------------------------------------------------
        # SHEET 3 — INVOICES
        # ---------------------------------------------------------
        ws_inv = wb.create_sheet(title="Invoices")
        ws_inv.views.sheetView[0].showGridLines = True
        
        inv_headers = ["Invoice ID", "Customer ID", "Customer Name", "Invoice Date", "Due Date", "Invoice Amount", "Currency"]
        for col, h in enumerate(inv_headers, 1):
            cell = ws_inv.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, inv in enumerate(invoices, start=2):
            row_vals = [
                inv.invoice_id, inv.customer_id, inv.customer_name,
                inv.invoice_date, inv.due_date, inv.invoice_amount, inv.currency
            ]
            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws_inv.cell(row=row_idx, column=col_idx, value=val)
                cell.font = regular_font
                cell.border = thin_border
                if col_idx == 6:
                    cell.number_format = '$#,##0.00'

        ws_inv.freeze_panes = 'A2'
        ws_inv.auto_filter.ref = ws_inv.dimensions

        # ---------------------------------------------------------
        # SHEET 4 — BANK TRANSACTIONS
        # ---------------------------------------------------------
        ws_txn = wb.create_sheet(title="Bank Transactions")
        ws_txn.views.sheetView[0].showGridLines = True
        
        txn_headers = ["Transaction ID", "Transaction Date", "Description", "Customer Name", "Reference", "Amount", "Currency"]
        for col, h in enumerate(txn_headers, 1):
            cell = ws_txn.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, txn in enumerate(transactions, start=2):
            row_vals = [
                txn.transaction_id, txn.transaction_date, txn.description,
                txn.customer_name, txn.reference or "", txn.amount, txn.currency
            ]
            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws_txn.cell(row=row_idx, column=col_idx, value=val)
                cell.font = regular_font
                cell.border = thin_border
                if col_idx == 6:
                    cell.number_format = '$#,##0.00'

        ws_txn.freeze_panes = 'A2'
        ws_txn.auto_filter.ref = ws_txn.dimensions

        # ---------------------------------------------------------
        # SHEET 5 — EXCEPTIONS
        # ---------------------------------------------------------
        ws_exc = wb.create_sheet(title="Exceptions")
        ws_exc.views.sheetView[0].showGridLines = True
        
        exc_sheet_headers = [
            "Transaction ID", "Transaction Date", "Customer Name", "Amount", "Currency",
            "Matched Invoice ID", "Match Type", "Confidence Level", "Payment Status",
            "Days Late", "Is Duplicate", "Action", "Exception Reason"
        ]
        
        for col, h in enumerate(exc_sheet_headers, 1):
            cell = ws_exc.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color="7F1D1D", end_color="7F1D1D", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        exceptions = [r for r in results if r.match_type == "unmatched" or r.action == "HUMAN_REVIEW" or r.is_duplicate or r.payment_status in ("late", "overdue")]

        for row_idx, r in enumerate(exceptions, start=2):
            reasons = []
            if r.match_type == "unmatched":
                reasons.append("Unmatched Transaction")
            if r.action == "HUMAN_REVIEW":
                reasons.append("Needs Human Review")
            if r.is_duplicate:
                reasons.append("Duplicate Payment Flag")
            if r.payment_status == "late":
                reasons.append(f"Late Payment ({r.days_late} days)")
            if r.payment_status == "overdue":
                reasons.append("Overdue Invoice")

            row_data = [
                r.transaction_id, r.transaction_date, r.customer_name, r.amount, r.currency,
                r.invoice_id or "", r.match_type, r.confidence_level, r.payment_status,
                r.days_late, "TRUE" if r.is_duplicate else "FALSE", r.action, ", ".join(reasons)
            ]

            for col_idx, val in enumerate(row_data, start=1):
                cell = ws_exc.cell(row=row_idx, column=col_idx, value=val)
                cell.font = regular_font
                cell.border = thin_border
                if col_idx == 4:
                    cell.number_format = '$#,##0.00'

        ws_exc.freeze_panes = 'A2'
        ws_exc.auto_filter.ref = ws_exc.dimensions

        # Auto-adjust column widths across all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if '\n' in val_str:
                        lines = val_str.split('\n')
                        max_len = max(max_len, max(len(l) for l in lines))
                    else:
                        max_len = max(max_len, len(val_str))
                sheet.column_dimensions[col_letter].width = max(12, min(max_len + 4, 40))

        stream = io.BytesIO()
        wb.save(stream)
        return stream.getvalue()
